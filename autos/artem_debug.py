from openpyxl.reader.excel import load_workbook
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from django.contrib.auth.models import User
from rest_framework import serializers

from autos.models import CollectableItem


def latitude_validator(value):
    if value < -90 or value > 90:
        raise serializers.ValidationError("Широта должна находиться в диапазоне от -90.0 до +90.0 градусов")
    return value


def longitude_validator(value):
    if value < -180 or value > 180:
        raise serializers.ValidationError("Долгота должна находиться в диапазоне от -180.0 до +180.0 градусов")
    return value


class CollectibleItemSerializer(serializers.ModelSerializer):
    url = serializers.URLField(source='picture')

    class Meta:
        model = CollectableItem
        exclude = ["picture"]
        extra_kwargs = {
            "latitude": {
                "validators": [latitude_validator],
            },
            "longitude": {
                "validators": [longitude_validator],
            },
        }


class FileUploadSerializer(serializers.Serializer):
    file = serializers.FileField(write_only=True, required=True)


class UploadFileView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        input_serializer = FileUploadSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        file_object = input_serializer.validated_data["file"]
        workbook = load_workbook(file_object)
        worksheet = workbook.active

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = list(h.lower() for h in header_row)

        wrong_rows = list()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            print(f'DEBUG {data}')
            row_serializer = CollectibleItemSerializer(data=data)
            if row_serializer.is_valid(raise_exception=False):
                row_serializer.save()
            else:
                print(f'DEBUG {",".join(wrong_rows)}')
                wrong_rows.append(list(row))

        return Response(wrong_rows, status=status.HTTP_200_OK)
